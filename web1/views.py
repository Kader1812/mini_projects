from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from .models import Brain
from.serializers import BrainSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import load_workbook
from openpyxl import Workbook
from io import BytesIO
from django.db import connection
import csv
import openpyxl
import MySQLdb



class InsertView(APIView):
   
  
   def post(self,request):
      
     # open the csv file
    with open('Brain.csv',mode='r')as file:
        
      #   serializers =BrainSerializer(data=request.data)
        

        # create a CSV reader objects
        reader = csv.DictReader(file)
        

        for row in reader:

            Brain.objects.create(
               Name = row['Name'],
               Production_link = row['Production link'],
               Description = row['Description']

            )
        
        return Response({'message':'store sucessfully'})


class CheckData(APIView):

    def post(self,request):
     
      csv_data =[]

      with open('Brain.csv',mode='r')as file:

        reader = csv.DictReader(file)
        
        for row in reader:
          
          product_link = row.get('Production link')
          description = row.get('Description')

          if product_link and description:
             csv_data.append({
                'product_link':product_link,
                'description':description
             })
          

        match_lst =[]
        not_match =[]

        for record in csv_data:
           
          product_link = record["product_link"]
          description = record["description"]

          if Brain.objects.filter(Production_link=product_link,Description=description).exists():
              
            match_lst.append(record)
          
          else:
             
            new_product = Brain.objects.create(Production_link=product_link,Description=description)
            not_match.append({
               
               'product_link':new_product.product_link,
               'description':new_product.description
            })
                
             

          
      return Response({'matching links':match_lst,'non match':not_match})


class GetData(APIView):
   
   def get(self,request):
      
      

      items = Brain.objects.all()

      serializer =BrainSerializer(items,many=True)

      return Response(serializer.data)

      



                   
              
           

           
           



        















# class CursorExportExcel(APIView):

#     def get(self,request):

#         with connection.cursor() as cursor:

#             cursor.execute(""" 
                          
            
#                                 SELECT ch.child_name,
#                                 ch.child_view_ID,
#                                 ch.child_id,
#                                 u.Users_FirstName AS enabler_name,
#                                 f.remarks,
#                                 f.feedback_id,
#                                 f.meeting_record_id,
#                                 f.class_readiness,
#                                 g.group_name,
#                                 imr.class_no,
#                                 attendance_meta.Meta_Desc AS attendance_description,
#                                 punctuality_meta.Meta_Desc AS punctuality_description,
#                                 learning_outcomes_meta.Meta_Desc AS learning_outcomes_description,
#                                 participation_level_meta.Meta_Desc AS participation_level_description,
#                                 involvement_meta.Meta_Desc AS involvement_description,
#                                 learning_enviornment_meta.Meta_Desc AS learning_enviornment_description,
#                                 internet_connectivity_meta.Meta_Desc AS internet_connectivity_description,
#                                 DATE(imr.scheduled_time) AS schedule_date
#                                 FROM 
#                                 3rlabs_child ch
#                                 JOIN 3rlabs_intervention_feedback f ON ch.child_id = f.child_id
#                                 JOIN 3rlabs_users u ON u.Users_id=f.enabler_id
#                                 JOIN 
#                                 3rlabs_meta AS attendance_meta
#                                 ON attendance_meta.Meta_id=f.attendance
#                                 JOIN
#                                 3rlabs_meta AS punctuality_meta
#                                 ON punctuality_meta.Meta_id=f.punctuality
#                                 JOIN
#                                 3rlabs_meta AS learning_outcomes_meta
#                                 ON learning_outcomes_meta.Meta_id=f.learning_outcomes
#                                 JOIN
#                                 3rlabs_meta AS participation_level_meta
#                                 ON participation_level_meta.Meta_id=f.participation_level
#                                 JOIN
#                                 3rlabs_meta AS involvement_meta
#                                 ON involvement_meta.Meta_id=f.involvement
#                                 JOIN
#                                 3rlabs_meta AS learning_enviornment_meta
#                                 ON learning_enviornment_meta.Meta_id=f.learning_enviornment
#                                 JOIN
#                                 3rlabs_meta AS internet_connectivity_meta
#                                 ON internet_connectivity_meta.Meta_id=f.internet_connectivity
#                                 JOIN
#                                 3rlabs_intervention_meeting_record imr ON f.meeting_record_id=imr.si_no
#                                 JOIN
#                                 3rlabs_intervention_meeting im ON imr.meeting_id=im.meeting_id
#                                 JOIN
#                                 3rlabs_intervention_group g ON im.group_id=g.group_id WHERE 
#                                 ch.child_view_ID = 'DL22PS298';
#              """)
            
#             intervention_rows = cursor.fetchall()
            
#             # Extract the column names for the result set

#             columns = [col[0] for col in cursor.description]

#             # Convert the fetched data into a list of dictionaries

#             intervention_data = [dict(zip(columns,row)) for row in intervention_rows]
            
#             # Initialize the final response data list
            
#             final_data = []
            
#             # Iterate over each intervention row and add assigned LWS and game information

#             for item in intervention_data:
                

#                 cursor.execute("""
#                                 SELECT 
#                                 imt.template_type,
#                                 img.game_name,
#                                 DATE_FORMAT(vch.created_on, '%d-%m-%Y') AS created_on
#                                 FROM 3rlabs_vakt_child_assign vch
#                                 JOIN 3rlabs_immersive_games img ON img.game_id = vch.game_id
#                                 JOIN 3rlabs_immersive_templates imt ON imt.t_id = img.template_id
#                                 WHERE Date_FORMAT(vch.created_on,'%Y-%m-%d') = '2022-08-10' 
#                                 AND vch.child_id = 1912;
#                 """

#                 )
                  
#                 results = cursor.fetchall()

#                 # Initialize empty lists to store 'lws' and 'game'

#                 lws_lst =[]
#                 game_lst = []

            
#                 for val in results:

#                     template_type = val[0]
#                     game_name =val[1]


#                     if template_type == 'lws':

#                         lws_lst.append(game_name)
#                     elif template_type == 'game':

#                         game_lst.append(game_name)
    
#                 #  convert list to a comma seperated string in python
               
#                 my_str_lws = ','.join(lws_lst)
#                 my_str_game = ','.join(game_lst)
                

#                 cursor.execute("""

#                         SELECT
#                         pw.name AS pws_name
#                         FROM pws_lessons pw
#                         JOIN 3rlabs_vakt_child_assign vch ON pw.si_no=vch.pws_session_id
#                         WHERE Date_FORMAT(vch.created_on,'%Y-%m-%d') = '2022-08-10'AND vch.child_id = 1912;
                        


#                 """ )

#                 pws_res = cursor.fetchall()

#                 pws_lst =[]

#                 for pws_items in pws_res:
                    
#                     pws_lst.append(pws_items[0])

#                 #convert list to a comma seperated string in python
                
#                 my_str_pws = ','.join(pws_lst)
                
                
#                 formatted_data = {
                    
#                     "child_name": item['child_name'],
#                     "child_view_ID": item['child_view_ID'],
#                     "child_id": item['child_id'],
#                     "enabler_name": item['enabler_name'],
#                     "remarks": item['remarks'],
#                     "feedback_id": item['feedback_id'],
#                     "meeting_record_id": item['meeting_record_id'],
#                     "class_readiness": item['class_readiness'],
#                     "group_name": item['group_name'],
#                     "class_no": item['class_no'],
#                     "attendance_description":item['attendance_description'],
#                     "punctuality_description":item['punctuality_description'],
#                     "learning_outcomes_description":item['learning_outcomes_description'],
#                     "participation_level_description":item['participation_level_description'],
#                     "involvement_description":item['involvement_description'],
#                     "learning_enviornment_description":item['learning_enviornment_description'],
#                     "internet_connectivity_description":item['internet_connectivity_description'],
#                     "schedule_date ":item['schedule_date'],
#                     "assigned_lws": my_str_lws,
#                     "assigned_game":my_str_game,
#                     "assigned_pws":my_str_pws

#                 }


#                 final_data.append(formatted_data)

#         # Creating a excel file
#         workbook = Workbook()
#         sheet = workbook.active
#         sheet.title = 'FinalImport'

#         # define the headers
#         headers = list(final_data[0].keys())

#         # import the headers in to excel

#         for col_num, header in enumerate(headers,1):

#             sheet.cell(row=1,column=col_num,value=header)
        
#         # define the rows

#         for row_num, data_item in enumerate(final_data,2):

#             for col_num, header in enumerate(headers,1):
                
#                 sheet.cell(row=row_num,column=col_num,value=data_item[header])

#         # Prepare response
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="FinalImport_data.xlsx"'
        
#         workbook.save(response)

#         return response
    


# class Production(APIView):

#     def get(self,request):

#         with connection.cursor() as cursor:

#             cursor.execute("""
            
#                         SELECT
#                         ch.child_name AS Child_Name,
#                         ptm.completed_class AS Class_no,
#                         DATE_FORMAT(ptm.created_on,'%d-%m-%Y') AS Date_Name,
#                         (CASE 
#                             WHEN ptm.enabler_remarks IS NULL THEN 'Incomplete'
#                             ELSE 'Complete'
#                         END) AS status_val,
#                         (		SELECT COUNT(imr.is_rescheduled_class) AS rescheduled_count
#                                 FROM 3rlabs_intervention_meeting_record imr 
#                                 JOIN 
#                                 3rlabs_intervention_meeting im ON imr.meeting_id=im.meeting_id
#                                 WHERE im.child_id =ch.child_id AND imr.scheduled_time<Date_Name AND imr.is_valid=1 AND imr.is_rescheduled=1
#                         )AS Reschedule_count,
#                         users.Users_FirstName AS Conducted_By,
#                         (		SELECT g_op.option
#                                 FROM ptm_analysis pta
#                                 JOIN 3rlabs_general_options g_op
#                                 ON pta.given_answer=g_op.si_no
#                                 WHERE pta.record_id=ptm.si_no AND pta.question_id=122
#                                 ORDER BY g_op.option Desc
#                                 LIMIT 1
#                         ) AS Preferred_renewal,
#                         ptm.enabler_remarks AS Enabler_remarks,
#                         ptm.parent_remarks AS Parent_remarks,
#                         ptm.internal_remarks AS Internal_remarks
#                         FROM ptm_meetings ptm
#                         JOIN
#                         3rlabs_users users ON users.Users_id=ptm.enabler_id
#                         JOIN
#                         3rlabs_child ch ON ch.child_id=ptm.child_id
#                         ORDER BY ptm.si_no DESC;

         



#             """)

#             meeting_rows = cursor.fetchall()


#             # creating a excel file
#             workbook = Workbook()
#             sheet = workbook.active
#             sheet.title = 'FinalImport2'


#         # initialize the headers

#         headers = ['Child_Name','Class_no','Date_Name','status_val','Reschedule_count',
#         'Conducted_By','Preferred_renewal','Enabler_remarks','Parent_remarks','Internal_remarks ']

#         # import the headers in the excel sheet
#         sheet.append(headers)


#         # import the rows in the worksheet

#         for row in meeting_rows:
            
#             sheet.append(row)

#         #  Create an HTTP response for downloading the Excel file
#             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename=Ptmdata.xlsx'


#         with BytesIO() as file_content:
#             workbook.save(file_content)
#             file_content.seek(0)
#             response.write(file_content.getvalue())
        
#         return response






    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
#         #     # creating a excel file
#         #     workbook = Workbook()
#         #     sheet = workbook.active
#         #     sheet.title = 'FinalImport2'


#         # # initialize the headers

#         # headers = ['Child_Name','Class_no','Date_Name','status_val','Reschedule_count',
#         # 'Conducted_By','Preferred_renewal','Enabler_remarks','Parent_remarks','Internal_remarks ']

#         # # import the headers in the excel sheet
#         # sheet.append(headers)


#         # # import the rows in the worksheet

#         # for row in meeting_rows:
            
#         #     sheet.append(row)

#         # #  Create an HTTP response for downloading the Excel file
#         #     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         #     response['Content-Disposition'] = 'attachment; filename=Ptmdata.xlsx'


#         # with BytesIO() as file_content:
#         #     workbook.save(file_content)
#         #     file_content.seek(0)
#         #     response.write(file_content.getvalue())
        
#         return JsonResponse(val,safe=False)
            




        
        
        
        
        
        
        
        
        
  
  
  
  
  
  
  
  
  
  
  
  
  
  
  
  
  
  
  
  

